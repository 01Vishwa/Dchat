import io
import re
import PyPDF2
import openpyxl
from docx import Document as DocxDocument
from typing import List, Dict


def _find_question_column(ws) -> int:
    """
    Detect which column contains the actual question text in an Excel sheet.
    Strategy:
      1. Check header row for keywords like 'question', 'description', 'text'.
      2. If no keyword match, pick the column with the longest average string values
         (the question column typically has the most text).
    Falls back to column 0.
    """
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value or '').strip().lower())

    # Priority keywords for a question column header
    keywords = ['question', 'description', 'text', 'query', 'item', 'detail', 'requirement']
    for kw in keywords:
        for idx, h in enumerate(headers):
            if kw in h:
                return idx

    # Fallback: pick column with longest average text length (skip pure-number columns)
    col_count = len(headers)
    if col_count <= 1:
        return 0

    avg_lengths: List[float] = []
    rows = list(ws.iter_rows(min_row=2, max_row=min(ws.max_row, 20)))  # sample first 20 rows
    for col_idx in range(col_count):
        total_len = 0
        str_count = 0
        for row in rows:
            if col_idx < len(row) and row[col_idx].value is not None:
                val = str(row[col_idx].value).strip()
                # Skip if the value is purely numeric
                try:
                    float(val)
                    continue
                except ValueError:
                    pass
                total_len += len(val)
                str_count += 1
        avg_lengths.append(total_len / max(str_count, 1))

    best_col = max(range(col_count), key=lambda i: avg_lengths[i])
    return best_col


def parse_questionnaire(file_bytes: bytes, filename: str) -> List[Dict[str, int | str]]:
    ext = filename.rsplit('.', 1)[-1].lower()
    questions = []
    
    if ext == 'pdf':
        reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
        text = " ".join(page.extract_text() for page in reader.pages)
        parts = re.split(r'\d+[\.\)]\s', text)
        questions = [p.strip() for p in parts if p.strip()]
        
    elif ext == 'xlsx':
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        q_col = _find_question_column(ws)
        for row in ws.iter_rows(min_row=2):
            if q_col < len(row) and row[q_col].value is not None:
                val = str(row[q_col].value).strip()
                if val:
                    questions.append(val)
        
    elif ext == 'docx':
        doc = DocxDocument(io.BytesIO(file_bytes))
        questions = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    
    else:
        raise ValueError(f"Unsupported file type: {ext}")
    
    return [{"number": i+1, "text": q} for i, q in enumerate(questions)]
